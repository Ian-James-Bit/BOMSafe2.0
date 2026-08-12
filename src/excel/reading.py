import pathlib
import openpyxl
from typing import Any
from dataclasses import dataclass

#so we can append API info next to relevant supplier column
@dataclass
class Supplier:
    name: str
    index: int

mpn_names = [
    "mpn",
    "manufacturerpartnumber",
    "mfrpartnumber",
    "manufacturerpart",
    "mfrpart",
    "manufacturerpartn",
    "mfrpartno",
    "manufacturerpn",
    "mfrpn",
    "partnumber",
]

supplier_names = [
    "supplier",
    "suppliers",
    "suppliername",
    "suppliersname",
    "distributor",
    "distrubutors",
    "distributors",
    "distributorname",
    "distributorsname",
]

# return a dict, each value of the dict is a mpn which is a dict holding row_number: int,
# suppliers: list of dataclass (name, column index)
def read_excel_file(path: pathlib.Path | str) -> dict[str,dict[str,int | Any]]:
    # Give the location of the file
    # path = "data/input/Test_Bom.xlsx"
    if(not isinstance(path, pathlib.Path | str)):
        raise ValueError("Path must be a path object or string.") 
    try:
        excel = openpyxl.load_workbook(path)
        #gets active sheet
        bom = excel.active

        row = bom.max_row
        mpn_index, supplier_indices = relevant_column_indices(bom)

        if(mpn_index is None):
            raise ValueError("MPN column not found in the Excel sheet.")
        if(len(supplier_indices) == 0):
            raise ValueError("Supplier columns not found in the Excel sheet.")
        # Process each row
        data_list: dict[str,dict[str,int | str | Any]] = {}
        for i in range(2, row + 1):
            mpn_value = bom.cell(row=i, column=mpn_index).value
            if ((mpn_value is None) | (not isinstance(mpn_value, str))):
                continue
            suppliers = []
            for supplier_index in supplier_indices:
                 supplier = Supplier(name = bom.cell(row=i, column=supplier_index).value,index = supplier_index)
                 if supplier.name is None:
                    continue
                 elif not isinstance(supplier.name, str):
                    continue
                 suppliers.append(supplier)
            # add it to list of relevant data
            data_list[mpn_value] = {"row_number": i,"suppliers": suppliers}

        return data_list
    except Exception as e:
            print(f"An unexpected error occured: {e}")
            raise
    finally:
         excel.close()
    
#bom represents a sheet
def relevant_column_indices(bom: Any) -> tuple[int | None, list[int]]:
    try:
        column = bom.max_column
        mpn_index: int | None = None
        supplier_indices: list[int] = []
        #look at first row for headers
        need_MPN=True
        for i in range(1, column + 1):
            cell_obj = bom.cell(row=1, column=i)
            if(need_MPN):
                if is_mpn_or_supplier(cell_obj.value,mpn_names):
                    mpn_index = i
                    need_MPN=False
                    continue
                
            if is_mpn_or_supplier(cell_obj.value,supplier_names):
                supplier_indices.append(i)

        return mpn_index, supplier_indices
    
    except Exception as e:
        print(f"An unexpected error occured: {e}")
        raise

#check if a header value is a mpn or supplier header
def is_mpn_or_supplier(name: str, aliases: list[str]) -> bool:
    # not a string therefore cant be an alias of "MPN" header
    if not isinstance(name, str):
        return False
    try:
        # get ride of characters that aren't letters
        temp = ""
        for character in name:
            if character.isalpha():
                temp += character
        name = temp.lower()
        for alias in aliases:
            if (alias == name):
                return True
        return False
    except Exception as e:
        print(f"An unexpected error occured: {e}")
        raise