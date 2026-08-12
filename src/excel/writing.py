# going to be a "convert_response" function in "API" folder that takes the response from
# whatever API we use and converts it into a dict of dicts, higher level dicts are by row 
# number (from bom dict) representing different mpns holding lower level dicts that are by 
#column index (from bom dict) representing different suppliers, holding stock, price,  ect..
from copy import copy
import pathlib
import openpyxl
from typing import Any
def add_data_to_bom(path: pathlib.Path | str, data: dict[int, dict[int | str, dict[Any, Any] | str]]) -> None:
    if(not isinstance(path, pathlib.Path | str)):
        raise ValueError("Path must be a path object or string.") 
    try:
        excel = openpyxl.load_workbook(path) 

        bom = excel.active 

        data = add_new_data_columns(bom, data)
        #mpn is row number
        for mpn in data:
            risk_cell = bom.cell(row=mpn, column=bom.max_column)
            risk_cell.value = data[mpn].get("risk")
            risk_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
            #suppplier is column index
            for supplier in data[mpn]:
                if isinstance(supplier, int):
                    #first new column for each supplier is stock, second is price
                    stock_cell = bom.cell(row=mpn, column=supplier+1)
                    price_cell = bom.cell(row=mpn, column=supplier+2)

                    stock_cell.value = data[mpn][supplier].get("stock")
                    price_cell.value = data[mpn][supplier].get("price")

                    stock_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
                    price_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        #fixing alignment of values of cells in the original bom sheet to be top aligned
        for row in bom.iter_rows(min_row=2):
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.vertical = "top"
                cell.alignment = alignment
        excel.save(path)
    except Exception as e:
        print(f"An unexpected error occured: {e}")
        raise
    finally:
        excel.close()

# creating new columns in the excel sheet next to the relevant supplier column if there is any data for that column in the response
# and updating the column indices in the data dict to reflect the new column indices in the excel sheet
def add_new_data_columns(bom, data: dict[int, dict[int | str, dict[Any, Any] | str]]) -> dict[int, dict[int | str, dict[Any, Any] | str]]:
    try:
        supplier_list = []
        # if there is a new unique supplier column_index in the data, add it to the list
        # because there is data for it that needs a column space to hold it in at least one row.
        # mpn is row index
        for mpn in data:
            # supplier is column index
            for supplier in data[mpn]:
                if isinstance(supplier, int): # only want to add new columns for supplier column indices, not the "risk" key
                    if supplier in supplier_list:
                        continue
                    else:
                        supplier_list.append(supplier)
        supplier_list.sort() # in case not in order        
        # make 3 new colums for each unique column index and update the column indices
        # depending on how many columns before it were added since everything is shifted now
        for column_index in supplier_list:
            total = 0
            for indices in supplier_list:
                if indices < column_index:
                    total += 1

            new_column_index = column_index + 2*total
            bom.insert_cols(new_column_index + 1,2)
            #name headers of the new columns
            bom.cell(row=1, column=new_column_index +1).value = "Stock"
            bom.cell(row=1, column=new_column_index + 2).value = "Price"
            #set column width to make it more readable
            bom.column_dimensions[openpyxl.utils.get_column_letter(new_column_index + 1)].width = 40
            bom.column_dimensions[openpyxl.utils.get_column_letter(new_column_index + 2)].width = 50
        # create column at the end of excel sheet for product risk and fill it in for each product (mpn)
        new_column = bom.max_column + 1
        bom.cell(row=1, column=new_column).value = "Product Risk"
        bom.column_dimensions[openpyxl.utils.get_column_letter(new_column)].width = 40
        # update the column indices in the data dict to reflect the new column indices in the excel sheet
        for mpn in data:
            new_mpn_value = {}
            new_mpn_value["risk"] = data[mpn].get("risk")
            for supplier in data[mpn]:
                if isinstance(supplier, int):
                    for column_index in supplier_list:
                        total = 0
                        for indices in supplier_list:
                            if indices < column_index:
                                total += 1
                        new_column_index = column_index + 2*total
                        if supplier == column_index:
                            new_mpn_value[new_column_index] = data[mpn][supplier]
            data[mpn] = (new_mpn_value)
        return data
    except Exception as e:
        print(f"An unexpected error occured: {e}")
        raise