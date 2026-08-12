import openpyxl
from src.excel import writing

def test_add_data_to_bom(test_excel_file, test_excel_file_6):
    for i in range(2):
        if i == 0:
            data = {
                2: {
                    "risk": "High",
                    2: {"stock": 10, "price": 5.99},
                },
                3: {
                    "risk": "Medium",
                    2: {"stock": 20, "price": 9.99},
                },
            }
            writing.add_data_to_bom(test_excel_file, data)
            file = test_excel_file
            workbook = openpyxl.load_workbook(file)
            bom = workbook.active
            assert bom.cell(row=2, column=3).value == 10
            assert bom.cell(row=2, column=4).value == 5.99
            assert bom.cell(row=2, column=bom.max_column).value == "High"
            assert bom.cell(row=3, column=3).value == 20
            assert bom.cell(row=3, column=4).value == 9.99
            assert bom.cell(row=3, column=bom.max_column).value == "Medium"
        else:
            data = {
                2: {
                    "risk": "High",
                    2: {"stock": 10, "price": 5.99},
                    4: {"stock": 15, "price": 7.99},
                },
                3: {
                    "risk": "Medium",
                    2: {"stock": 20, "price": 9.99},
                    4: {"stock": 25, "price": 12.99},
                },
            }
            writing.add_data_to_bom(test_excel_file_6, data)
            file = test_excel_file_6
            workbook = openpyxl.load_workbook(file)
            bom = workbook.active
            assert bom.cell(row=2, column=3).value == 10
            assert bom.cell(row=2, column=4).value == 5.99
            assert bom.cell(row=2, column=7).value == 15
            assert bom.cell(row=2, column=8).value == 7.99
            assert bom.cell(row=2, column=bom.max_column).value == "High"
            assert bom.cell(row=3, column=3).value == 20
            assert bom.cell(row=3, column=4).value == 9.99
            assert bom.cell(row=3, column=7).value == 25
            assert bom.cell(row=3, column=8).value == 12.99
            assert bom.cell(row=3, column=bom.max_column).value == "Medium"

def test_add_new_data_columns(test_excel_file, test_excel_file_6):
    for i in range(2):
        if i == 0:
            file = test_excel_file
            workbook = openpyxl.load_workbook(file)
            bom = workbook.active

            # Create a sample data dictionary to add new columns
            data = {
                2: {2: {"risk": "High", "stock": 10, "price": 5.99}},
                3: {2: {"risk": "Medium", "stock": 20, "price": 9.99}},
            }

            updated_data = writing.add_new_data_columns(bom, data)

            # Check if new columns were added and headers are correct
            assert bom.cell(row=1, column=3).value == "Stock"
            assert bom.cell(row=1, column=4).value == "Price"
            assert bom.cell(row=1, column=bom.max_column).value == "Product Risk"
            
            #check column indeces are same
            for row in data:
                for key in data[row]:
                    if isinstance(key, int):
                        for updated_key in updated_data[row]:
                            if isinstance(updated_key, int):
                                assert key == updated_key
        else:
            for i in range(2):
                file = test_excel_file_6
                workbook = openpyxl.load_workbook(file)
                bom = workbook.active

                if i == 0:
                    data = {
                        2: {"risk": "High", 2: {"stock": 10, "price": 5.99}, 4: {"stock": 15, "price": 7.99}},
                        3: {"risk": "Medium", 2: {"stock": 20, "price": 9.99}, 4: {"stock": 25, "price": 12.99}},
                    }

                    updated_data = writing.add_new_data_columns(bom, data)

                    # Check if new columns were added and headers are correct
                    assert bom.cell(row=1, column=3).value == "Stock"
                    assert bom.cell(row=1, column=4).value == "Price"
                    assert bom.cell(row=1, column=7).value == "Stock"
                    assert bom.cell(row=1, column=8).value == "Price"
                    assert bom.cell(row=1, column=bom.max_column).value == "Product Risk"

                    #check if column indeces are updated correctly
                    for row in updated_data:
                        for key in data[row]:
                            if isinstance(key, int):
                                if key == 4:
                                    if isinstance(data[row][key], int):
                                        assert key == data[row][key]+2
                else:
                    #if no data was found of a specific supplier before another supplier, then the column index should be the same as before
                    data = {
                        2: {4: {"risk": "High", "stock": 15, "price": 7.99, }},
                        3: {4: {"risk": "Medium", "stock": 25, "price": 12.99 }},
                    }

                    updated_data = writing.add_new_data_columns(bom, data)
                    
                    #check if new columns are headers are correct
                    assert bom.cell(row=1, column=5).value == "Stock"
                    assert bom.cell(row=1, column=6).value == "Price"
                    assert bom.cell(row=1, column=bom.max_column).value == "Product Risk"
                    #check column indeces are same
                    for row in data:
                        for key in data[row]:
                            if isinstance(key, int):
                                for updated_key in updated_data[row]:
                                    if isinstance(data[row][key], int):
                                        assert key == updated_key
