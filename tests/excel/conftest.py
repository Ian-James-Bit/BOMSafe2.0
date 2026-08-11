import pytest
import openpyxl
from pathlib import Path
#test excel files for reading and writing functions with different column names and data types

@pytest.fixture
def test_excel_file():
    # Create a temporary Excel file for testing
    test_file_path = Path(__file__).parent / "test_excel.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    sheet.append(["MPN", "Supplier", "Description"])
    sheet.append(["12345", "Supplier A", "Test Description 1"])
    sheet.append(["67890", "Supplier B", "Test Description 2"])
    workbook.save(test_file_path)
    yield test_file_path
    test_file_path.unlink()  # Clean up the temporary file after the test

@pytest.fixture
def test_excel_file_2():
    # Create a temporary Excel file for testing
    test_file_path = Path(__file__).parent / "test_2_excel.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    sheet.append(["MPN", "Random Name", "Description"])
    sheet.append([12345, "Supplier A", "Test Description 1"])
    sheet.append([67890, "Supplier B", "Test Description 2"])
    workbook.save(test_file_path)
    yield test_file_path
    test_file_path.unlink()  # Clean up the temporary file after the test

@pytest.fixture
def test_excel_file_3():
    # Create a temporary Excel file for testing
    test_file_path = Path(__file__).parent / "test_3_excel.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    sheet.append(["Random Number", "Supplier", "Description"])
    sheet.append([12345, "Supplier A", "Test Description 1"])
    sheet.append([67890, "Supplier B", "Test Description 2"])
    workbook.save(test_file_path)
    yield test_file_path
    test_file_path.unlink()  # Clean up the temporary file after the test

@pytest.fixture
def test_excel_file_4():
    # Create a temporary Excel file for testing
    test_file_path = Path(__file__).parent / "test_4_excel.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    sheet.append(["Random Number", "Random Name", "Description"])
    sheet.append([12345, "Supplier A", "Test Description 1"])
    sheet.append([67890, "Supplier B", "Test Description 2"])
    workbook.save(test_file_path)
    yield test_file_path
    test_file_path.unlink()  # Clean up the temporary file after the test

@pytest.fixture
def test_excel_file_5():
    # Create a temporary Excel file for testing
    test_file_path = Path(__file__).parent / "test_5_excel.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    sheet.append(["MPN", "Supplier", "Description"])
    sheet.append([None, None, "Test Description 1"])
    sheet.append([None, 434, "Test Description 2"])
    workbook.save(test_file_path)
    yield test_file_path
    test_file_path.unlink()  # Clean up the temporary file after the test

@pytest.fixture
def test_excel_file_6():
    # Create a temporary Excel file for testing
    test_file_path = Path(__file__).parent / "test_6_excel.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    sheet.append(["MPN", "Supplier", "Description", "Supplier 2"])
    sheet.append(["12345", "Supplier A", "Test Description 1", "Supplier B"])
    sheet.append(["67890", "Supplier B", "Test Description 2", "Supplier C"])
    workbook.save(test_file_path)
    yield test_file_path
    test_file_path.unlink()  # Clean up the temporary file after the test